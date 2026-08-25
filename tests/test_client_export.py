from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import AsyncMock

import pytest
from openpyxl import load_workbook
from telegram.ext import ConversationHandler

from app import storage
from app.config import TZ
from app.subscriptions.requests import operations as product_operations
from app.users.admin.client_export import CLIENT_HEADERS, REQUEST_HEADERS, build_client_workbook
from app.users.admin.export_handlers import export_clients_xlsx_cb
from tests.product_support import _admin, _callback_update, _user


def test_workbook_contains_clients_and_request_history_without_connection_secrets() -> None:
    generated_at = datetime(2026, 8, 11, 12, 30, tzinfo=TZ)
    owner = _admin(1, admin_level="owner")
    client = _user(
        42,
        nickname='=HYPERLINK("https://evil.test")',
        username="formula_user",
        contact_email="client@example.com",
        connection_url="https://connect.test/private-token",
        service_tier="subscriber",
        is_paid=True,
        paid_at=(generated_at - timedelta(days=30)).isoformat(),
        subscription_end_at=(generated_at + timedelta(days=60)).isoformat(),
    )
    config = storage.UserData(authorized_users={"1": owner, "42": client})
    request = product_operations.create_request(
        config,
        kind="purchase",
        user_id=42,
        status="approved",
        target_end_at=(generated_at + timedelta(days=60)).isoformat(),
    )

    content, client_count, request_count = build_client_workbook(
        config.authorized_users,
        config.service_requests,
        generated_at=generated_at,
    )

    assert (client_count, request_count) == (1, 1)
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)
    assert workbook.sheetnames == ["Клиенты", "Заявки"]
    clients = workbook["Клиенты"]
    requests = workbook["Заявки"]
    assert tuple(cell.value for cell in clients[1]) == CLIENT_HEADERS
    assert tuple(cell.value for cell in requests[1]) == REQUEST_HEADERS
    client_row = {
        clients.cell(1, column).value: clients.cell(2, column).value for column in range(1, clients.max_column + 1)
    }
    request_row = {
        requests.cell(1, column).value: requests.cell(2, column).value for column in range(1, requests.max_column + 1)
    }
    assert client_row["Telegram ID"] == 42
    assert client_row["Никнейм"].startswith("'=")
    assert client_row["Оплачен"] == "Да"
    assert client_row["Ссылка настроена"] == "Да"
    assert request_row["ID заявки"] == request["id"]
    assert request_row["Тип"] == "Покупка"
    all_values = "\n".join(
        str(cell.value or "") for worksheet in workbook.worksheets for row in worksheet.iter_rows() for cell in row
    )
    assert "private-token" not in all_values
    workbook.close()


@pytest.mark.asyncio
async def test_owner_export_handler_sends_xlsx_and_audits_action(isolated_storage: None) -> None:
    def _seed(config: storage.UserData) -> None:
        config.authorized_users = {
            "1": _admin(1, admin_level="owner"),
            "42": _user(42),
        }
        product_operations.create_request(config, kind="trial", user_id=42)

    await storage.update_user_data(_seed)
    update, context = _callback_update(1, "users:export:xlsx")
    context.bot = SimpleNamespace(send_document=AsyncMock())

    await export_clients_xlsx_cb(update, context)

    context.bot.send_document.assert_awaited_once()
    document = context.bot.send_document.await_args.kwargs["document"]
    assert document.filename.startswith("maintbot_clients_")
    assert document.filename.endswith(".xlsx")
    workbook = load_workbook(BytesIO(document.input_file_content), read_only=True)
    assert workbook["Клиенты"].max_row == 2
    assert workbook["Заявки"].max_row == 2
    workbook.close()
    assert storage.audit_log_snapshot()[-1]["action"] == "clients_exported"


@pytest.mark.asyncio
async def test_non_owner_cannot_export_clients(isolated_storage: None) -> None:
    await storage.update_user_data(lambda config: config.authorized_users.update({"2": _admin(2), "42": _user(42)}))
    update, context = _callback_update(2, "users:export:xlsx")
    context.bot = SimpleNamespace(send_document=AsyncMock())

    result = await export_clients_xlsx_cb(update, context)

    assert result == ConversationHandler.END
    context.bot.send_document.assert_not_awaited()
    update.effective_message.reply_text.assert_awaited_once_with("Это действие доступно только руководителю сервиса.")
