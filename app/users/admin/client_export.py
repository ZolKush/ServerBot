"""Build owner-only spreadsheet exports of client and request data."""

from __future__ import annotations

import re
from collections.abc import Iterable, Mapping, Sequence
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_ILLEGAL_XML_CHARACTERS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_FORMULA_PREFIXES = ("=", "+", "-", "@")

_ACCESS_LABELS = {
    "approved": "Одобрен",
    "pending": "Ожидает решения",
    "rejected": "Отклонён",
    "blocked": "Заблокирован",
    "logged_out": "Вышел",
}
_TIER_LABELS = {
    "basic": "Базовый",
    "subscriber": "Подписчик",
    "unlimited_trial": "Бессрочный тест",
}
_REQUEST_KIND_LABELS = {
    "trial": "Тестовый доступ",
    "purchase": "Покупка",
    "renewal": "Продление",
}
_REQUEST_STATUS_LABELS = {
    "pending": "Ожидает решения",
    "claimed": "В обработке",
    "awaiting_link": "Ожидает ссылку",
    "requisites_sent": "Реквизиты отправлены",
    "payment_reported": "Оплата заявлена",
    "approved": "Одобрена",
    "rejected": "Отклонена",
    "cancelled": "Отменена",
}

CLIENT_HEADERS = (
    "Telegram ID",
    "Никнейм",
    "Username",
    "Имя",
    "Фамилия",
    "Email",
    "Статус доступа",
    "Активен",
    "Уровень сервиса",
    "Оплачен",
    "Оплата подтверждена",
    "Подписка до",
    "Подтвердил оплату",
    "Тест выдан",
    "Тест до",
    "Тест, часов",
    "Ссылка настроена",
    "Первая авторизация",
    "Запрос доступа",
    "Решение по доступу",
    "Выход",
    "Обновление подписки",
    "Последнее автонапоминание",
    "Последнее ручное напоминание",
)

REQUEST_HEADERS = (
    "ID заявки",
    "Telegram ID",
    "Клиент",
    "Тип",
    "Статус",
    "Создана",
    "Обновлена",
    "Целевая дата",
    "Оплата заявлена",
    "Рассмотрена",
    "Рассмотрел, ID",
    "Тест, часов",
    "Причина решения",
    "Комментарий",
)


def _safe_text(value: object) -> str:
    """Keep arbitrary profile data as inert worksheet text."""

    text = _ILLEGAL_XML_CHARACTERS.sub("", str(value or ""))
    if text.lstrip().startswith(_FORMULA_PREFIXES):
        return "'" + text
    return text


def _yes_no(value: object) -> str:
    return "Да" if bool(value) else "Нет"


def _display_name(meta: Mapping[str, Any]) -> str:
    nickname = _safe_text(meta.get("nickname")).strip()
    if nickname:
        return nickname
    username = _safe_text(meta.get("username")).strip().lstrip("@")
    if username:
        return f"@{username}"
    name = " ".join(
        part for part in (_safe_text(meta.get("first_name")).strip(), _safe_text(meta.get("last_name")).strip()) if part
    )
    return name or _safe_text(meta.get("user_id"))


def _client_rows(users: Mapping[str, Mapping[str, Any]]) -> Iterable[Sequence[object]]:
    clients = [meta for meta in users.values() if isinstance(meta, Mapping) and meta.get("role") != "admin"]
    clients.sort(key=lambda meta: int(meta.get("user_id") or 0))
    for meta in clients:
        access_state = str(meta.get("access_state") or "")
        tier = str(meta.get("service_tier") or "")
        yield (
            int(meta.get("user_id") or 0),
            _safe_text(meta.get("nickname")),
            _safe_text(meta.get("username")),
            _safe_text(meta.get("first_name")),
            _safe_text(meta.get("last_name")),
            _safe_text(meta.get("contact_email")),
            _ACCESS_LABELS.get(access_state, _safe_text(access_state)),
            _yes_no(meta.get("enabled")),
            _TIER_LABELS.get(tier, _safe_text(tier)),
            _yes_no(meta.get("is_paid")),
            _safe_text(meta.get("paid_at")),
            _safe_text(meta.get("subscription_end_at")),
            _safe_text(meta.get("payment_confirmed_by_name") or meta.get("payment_confirmed_by_id")),
            _safe_text(meta.get("trial_issued_at")),
            _safe_text(meta.get("trial_end_at")),
            meta.get("trial_duration_hours") or "",
            _yes_no(bool(str(meta.get("connection_url") or "").strip())),
            _safe_text(meta.get("auth_at")),
            _safe_text(meta.get("access_requested_at")),
            _safe_text(meta.get("access_reviewed_at")),
            _safe_text(meta.get("logged_out_at")),
            _safe_text(meta.get("subscription_updated_at")),
            _safe_text(meta.get("last_auto_payment_reminder_at")),
            _safe_text(meta.get("last_manual_payment_reminder_at")),
        )


def _request_rows(
    requests: Mapping[str, Mapping[str, Any]],
    users: Mapping[str, Mapping[str, Any]],
) -> Iterable[Sequence[object]]:
    items = [request for request in requests.values() if isinstance(request, Mapping)]
    items.sort(key=lambda request: int(request.get("id") or 0))
    for request in items:
        user_id = int(request.get("user_id") or 0)
        meta = users.get(str(user_id)) or {}
        if meta.get("role") == "admin":
            continue
        kind = str(request.get("kind") or "")
        status = str(request.get("status") or "")
        yield (
            int(request.get("id") or 0),
            user_id,
            _display_name(meta),
            _REQUEST_KIND_LABELS.get(kind, _safe_text(kind)),
            _REQUEST_STATUS_LABELS.get(status, _safe_text(status)),
            _safe_text(request.get("created_at")),
            _safe_text(request.get("updated_at")),
            _safe_text(request.get("target_end_at")),
            _safe_text(request.get("payment_reported_at")),
            _safe_text(request.get("reviewed_at")),
            request.get("reviewed_by_id") or "",
            request.get("trial_duration_hours") or "",
            _safe_text(request.get("decision_reason")),
            _safe_text(request.get("comment")),
        )


def _populate_sheet(
    worksheet: Any,
    *,
    headers: Sequence[str],
    rows: Iterable[Sequence[object]],
) -> int:
    worksheet.append(tuple(headers))
    row_count = 0
    for row in rows:
        worksheet.append(tuple(row))
        row_count += 1

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, row_count + 1)}"
    worksheet.row_dimensions[1].height = 32

    for column_index, header in enumerate(headers, start=1):
        values = [header]
        values.extend(
            str(worksheet.cell(row=row_index, column=column_index).value or "") for row_index in range(2, row_count + 2)
        )
        width = min(45, max(12, max(len(value) for value in values) + 2))
        worksheet.column_dimensions[get_column_letter(column_index)].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    return row_count


def build_client_workbook(
    users: Mapping[str, Mapping[str, Any]],
    requests: Mapping[str, Mapping[str, Any]],
    *,
    generated_at: datetime,
) -> tuple[bytes, int, int]:
    """Return XLSX bytes plus exported client and request counts."""

    workbook = Workbook()
    client_sheet = workbook.active
    if client_sheet is None:
        raise RuntimeError("workbook has no active worksheet")
    client_sheet.title = "Клиенты"
    client_count = _populate_sheet(client_sheet, headers=CLIENT_HEADERS, rows=_client_rows(users))

    request_sheet = workbook.create_sheet("Заявки")
    request_count = _populate_sheet(
        request_sheet,
        headers=REQUEST_HEADERS,
        rows=_request_rows(requests, users),
    )

    workbook.properties.title = "MaintBot — выгрузка клиентов"
    workbook.properties.subject = f"Сформировано {generated_at.isoformat()}"
    workbook.properties.creator = "MaintBot"
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue(), client_count, request_count


def client_export_filename(generated_at: datetime) -> str:
    return f"maintbot_clients_{generated_at:%Y%m%d_%H%M%S}.xlsx"


__all__ = [
    "CLIENT_HEADERS",
    "REQUEST_HEADERS",
    "build_client_workbook",
    "client_export_filename",
]
